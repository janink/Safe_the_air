import openpyxl

excelFile = 'Neighboors.xlsx'
data = openpyxl.load_workbook(excelFile)
old = data.get_sheet_by_name('old')
new = data.get_sheet_by_name('new')

columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN']
for i in range(2, old.max_row+1):
    j = 0
    while old[str(columns[j])+str(i)].value != None:
        new['A'+str(new.max_row+1)] = old['A'+str(i)].value
        new['B'+str(new.max_row)] = old[str(columns[j])+str(i)].value
        j += 1

data.save(excelFile)