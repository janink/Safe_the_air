import openpyxl
import pickle
from scipy.spatial import distance
import math

excelFile = 'LondonStreetData.xlsx'
data = openpyxl.load_workbook(excelFile)
london = data.get_sheet_by_name('london')

excelFile = 'LondonStreetNeighboorData.xlsx'
data = openpyxl.load_workbook(excelFile)
neighboors = data.get_sheet_by_name('neighboors')

class London:
    def __init__(self, streetID, streetName, area, roadStatus, neighboors, lat, long):
        self.streetID = streetID
        self.streetName = streetName
        self.area = area
        self.roadStatus = roadStatus
        self.neighboors = neighboors
        self.lat = lat
        self.long = long

class Roadstatus:
    def __init__(self, color):
        self.color = color

class Neighboor:
    def __init__(self, hasNeighboor):
        self.hasNeighboor = hasNeighboor

'''        
def getNeighboors(sheet):
    try:
        with open("pickle/neighboorsDictionary.txt", "rb") as myFile:
            neighboorsDictionary = pickle.load(myFile)
    except:
        neighboorsDictionary, neighboorsList = {}, []
        for i in range(2, sheet.max_row+1):
            neighboorsList.append(sheet['B'+str(i)].value)
            if sheet['A'+str(i)].value != sheet['A'+str(i+1)].value:
                neighboorsDictionary[str(sheet['A'+str(i)].value)] =  Neighboor(neighboorsList)
                neighboorsList = []
        with open("pickle/neighboorsDictionary.txt", "wb") as myFile:
            pickle.dump(neighboorsDictionary, myFile)
    return neighboorsDictionary
'''

def getLondonData(sheet):#, neighboorsDictionary):
    try:
        with open("pickle/londonDataDictionary.txt", "rb") as myFile:
            londonDataDictionary = pickle.load(myFile)
    except:
        londonDataDictionary = {}
        for i in range(2, sheet.max_row+1):
            streetID = sheet['A'+str(i)].value
            streetName = sheet['D'+str(i)].value
            area = sheet['G'+str(i)].value
            roadStatus = sheet['H'+str(i)].value
            neighboors = 1#neighboorsDictionary[str(sheet['A'+str(i)].value)]
            lat = sheet['E'+str(i)].value
            long = sheet['F'+str(i)].value
            londonDataDictionary[str(sheet['A'+str(i)].value)] = London(streetID, streetName, area, roadStatus, neighboors, lat, long)      
        with open("pickle/londonDataDictionary.txt", "wb") as myFile:
            pickle.dump(londonDataDictionary, myFile)
    return londonDataDictionary

def getRoadStatus(sheet):
    try:
        with open("pickle/roadStatusDictionary.txt", "rb") as myFile:
            roadStatusDictionary = pickle.load(myFile)
    except:
        roadStatusDictionary = {}
        listOptions = [None, 'G', 'Y', 'R']
        for i in range(2, sheet.max_row+1):
            roadStatusDictionary[str(sheet['A'+str(i)].value)] = Roadstatus(listOptions[sheet['H'+str(i)].value])
        with open("pickle/roadStatusDictionary.txt", "wb") as myFile:
            pickle.dump(roadStatusDictionary, myFile)
    return roadStatusDictionary

def getClosestNeighboor(neighboors, point, goalLat, goalLong, londonData):
    closestNeighboor = point
    closestDistance = math.inf
    for i in range(len(neighboors)):
        a = (londonData[neighboors[i]].lat, goalLat)
        b = (londonData[neighboors[i]].long, goalLong)
        dst = distance.euclidean(a,b)
        if closestDistance > dst:
            closestDistance = dst
            closestNeighboor = neighboors[i]
    return closestNeighboor   
    
def findshortestPath(start, goal, londonData):
    pathList = [start]
    goalLat = londonData[goal].lat
    goalLong = londonData[goal].long
    point = start
    while point != goal:
        if goal in londonData[point].neighboors:
            break
        closestNeighboor = getClosestNeighboor(londonData[point].neighboors, point, goalLat, goalLong, londonData)      
        pathList.append(closestNeighboor)
        point = closestNeighboor
    pathList.append(goal)
    return pathList

def getClosestNeighboorWithErrorList(neighboors, point, goalLat, goalLong, londonData, errorList):
    closestNeighboor = point
    closestDistance = math.inf
    for i in range(len(neighboors)):
        if neighboors[i] in errorList:
            continue
        a = (londonData[neighboors[i]].lat, goalLat)
        b = (londonData[neighboors[i]].long, goalLong)
        dst = distance.euclidean(a,b)
        if closestDistance > dst:
            closestDistance = dst
            closestNeighboor = neighboors[i]
    return closestNeighboor   

def findAlternativePathPart(start, goal, londonData):
    pathList = [start]
    goalLat = londonData[goal].lat
    goalLong = londonData[goal].long
    point = start
    while point != goal:
        if goal in londonData[point].neighboors:
            break
        closestNeighboor = getClosestNeighboor(londonData[point].neighboors, point, goalLat, goalLong, londonData)      
        pathList.append(closestNeighboor)
        point = closestNeighboor
    pathList.append(goal)
    return pathList

def findPathEnvironment(start, goal, londonData, shortestPath):
    environmentalPath = shortestPath
    StillRed = True
    errorlist = []
    while StillRed == True:
        for i in range(len(environmentalPath)):
            if environmentalPath[i] == red:
                StillRed == True
                break
            StillRed = False
        errorlist.append(environmentalPath[i])
        environmentalPath = environmentalPath[0:i]
        pathRest = findAlternativePathPart(start, goal, londonData)
        for j in range(len(pathRest)):
            environmentalPath.appnd(pathRest[j])
    environmentalPath.append(goal)
    return environmentalPath


#-------------------------run---------------------------------------

#neighboorsDictionary = getNeighboors(neighboors)
londonDataDictionary = getLondonData(london), #neighboorsDictionary)
roadStatusDictionary = getRoadStatus(london)

start = '331357'
goal = '8653432'

shortestPath = findshortestPath(start, goal, londonDataDictionary)
environmentPath = findPathEnvironment(londonDataDictionary, shortestPath)
